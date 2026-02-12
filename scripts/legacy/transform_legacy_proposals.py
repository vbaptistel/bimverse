#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel
except ModuleNotFoundError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: python3 -m pip install --user openpyxl"
    ) from exc

SKIP_SHEETS = {"RESUMO"}
PROPOSAL_CODE_RE = re.compile(r"^BV-([A-Z0-9]+)-(\d{4})-BIM-(\d{4})$")


@dataclass
class CustomerRecord:
    name: str
    slug: str
    status: str = "ativa"
    cnpj: str = ""
    notes: str = ""


@dataclass
class ProposalRecord:
    customer_slug: str
    code: str
    seq_number: int
    year: int
    invitation_code: str
    project_name: str
    scope_description: str
    status: str
    estimated_value_brl: str
    final_value_brl: str
    outcome_reason: str
    created_at: str
    updated_at: str
    legacy_sheet: str
    legacy_row: int


@dataclass
class ProposalRevisionRecord:
    proposal_code: str
    revision_number: int
    value_before_brl: str
    value_after_brl: str
    reason: str
    scope_changes: str
    discount_brl: str
    discount_percent: str
    notes: str
    created_at: str
    legacy_sheet: str
    legacy_row: int
    legacy_revision_label: str


@dataclass
class SheetMapping:
    description_col: int | None
    invitation_col: int | None
    active_col: int | None
    won_col: int | None
    lost_col: int | None
    rev_value_cols: dict[int, int]
    data_cols: list[int]
    headers: dict[int, str]


def normalize_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def normalize_header(value: object) -> str:
    text = normalize_str(value)
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"\s+", " ", text)
    return text.strip().upper()


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_only = ascii_only.lower()
    ascii_only = re.sub(r"[^a-z0-9]+", "-", ascii_only)
    ascii_only = ascii_only.strip("-")
    return ascii_only or "cliente"


def unique_slug(base: str, used: set[str]) -> str:
    if base not in used:
        used.add(base)
        return base

    index = 2
    while True:
        candidate = f"{base}-{index}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        index += 1


def parse_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return None
    if isinstance(value, date):
        return None

    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    raw = normalize_str(value)
    if not raw:
        return None

    raw = raw.replace("R$", "").replace(" ", "")

    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(".", "").replace(",", ".")

    try:
        return Decimal(raw).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return None


def parse_date_value(value: object, *, allow_excel_serial: bool) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if allow_excel_serial and isinstance(value, (int, float)):
        if 20000 <= float(value) <= 80000:
            try:
                converted = from_excel(value)
                if isinstance(converted, datetime):
                    return converted.date()
                if isinstance(converted, date):
                    return converted
            except Exception:
                return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue

    return None


def decimal_to_csv(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}"


def date_to_timestamp_csv(value: date) -> str:
    return f"{value.isoformat()}T00:00:00Z"


def pick_first(*values: Decimal | None) -> Decimal | None:
    for value in values:
        if value is not None:
            return value
    return None


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Transforms legacy proposals spreadsheet into CSV and SQL import files."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to legacy .xlsx file",
    )
    parser.add_argument(
        "--output-dir",
        default="tmp/legacy-import",
        help="Output directory (default: tmp/legacy-import)",
    )
    return parser


def extract_sheet_mapping(worksheet) -> SheetMapping:
    headers: dict[int, str] = {}
    for col in range(1, worksheet.max_column + 1):
        normalized = normalize_header(worksheet.cell(row=4, column=col).value)
        if normalized:
            headers[col] = normalized

    def find_first(pattern: str) -> int | None:
        regex = re.compile(pattern)
        for col in sorted(headers):
            if regex.search(headers[col]):
                return col
        return None

    rev_value_cols: dict[int, int] = {}
    for revision in (0, 1, 2):
        pattern = rf"TOTAL\s*REV\.?\s*{revision}"
        col = find_first(pattern)
        if col is not None:
            rev_value_cols[revision] = col

    data_cols = [col for col, header in headers.items() if header == "DATA"]

    return SheetMapping(
        description_col=find_first(r"DESCRI"),
        invitation_col=find_first(r"CARTA\s+CONVITE"),
        active_col=find_first(r"TOTAL\s+REV\s+ATIVA\s+EM\s+CONCORRENCIA"),
        won_col=find_first(r"GANHOU\s+CONCORRENCIA"),
        lost_col=find_first(r"PERDEU\s+CONCORRENCIA"),
        rev_value_cols=rev_value_cols,
        data_cols=data_cols,
        headers=headers,
    )


def parse_revisions_from_row(worksheet, row: int, mapping: SheetMapping) -> list[dict[str, object]]:
    parsed: list[dict[str, object]] = []
    used_date_cols: set[int] = set()

    all_rev_cols = sorted(mapping.rev_value_cols.values())

    for revision_label in sorted(mapping.rev_value_cols):
        value_col = mapping.rev_value_cols[revision_label]
        raw_value = worksheet.cell(row=row, column=value_col).value
        revision_value = parse_decimal(raw_value)

        next_bound_candidates = [col for col in all_rev_cols if col > value_col]
        if mapping.won_col is not None:
            next_bound_candidates.append(mapping.won_col)
        if mapping.lost_col is not None:
            next_bound_candidates.append(mapping.lost_col)
        next_bound = min(next_bound_candidates) if next_bound_candidates else worksheet.max_column + 1

        revision_date: date | None = None
        explicit_data_cols = [col for col in mapping.data_cols if value_col < col < next_bound]
        if explicit_data_cols:
            selected_data_col = explicit_data_cols[0]
            revision_date = parse_date_value(
                worksheet.cell(row=row, column=selected_data_col).value,
                allow_excel_serial=True,
            )
            if revision_date is not None:
                used_date_cols.add(selected_data_col)

        if revision_value is None:
            continue

        parsed.append(
            {
                "label": revision_label,
                "value_col": value_col,
                "value": revision_value,
                "date": revision_date,
            }
        )

    if not parsed:
        return parsed

    # If date is still missing, try immediate right column.
    for revision in parsed:
        if revision["date"] is not None:
            continue

        value_col = int(revision["value_col"])
        right_col = value_col + 1
        right_header = mapping.headers.get(right_col, "")
        allow_serial = right_header == "DATA"

        inferred_date = parse_date_value(
            worksheet.cell(row=row, column=right_col).value,
            allow_excel_serial=allow_serial,
        )
        if inferred_date is not None:
            revision["date"] = inferred_date
            used_date_cols.add(right_col)

    # Fallback: capture orphan date-like cells between first revision and outcome columns.
    region_start = min(int(item["value_col"]) for item in parsed)
    outcome_candidates = [col for col in (mapping.won_col, mapping.lost_col) if col is not None]
    region_end = min(outcome_candidates) - 1 if outcome_candidates else worksheet.max_column

    value_cols = {int(item["value_col"]) for item in parsed}
    for col in range(region_start, region_end + 1):
        if col in value_cols or col in used_date_cols:
            continue

        allow_serial = mapping.headers.get(col, "") == "DATA"
        orphan_date = parse_date_value(
            worksheet.cell(row=row, column=col).value,
            allow_excel_serial=allow_serial,
        )
        if orphan_date is None:
            continue

        candidates = [
            item
            for item in parsed
            if item["date"] is None and int(item["value_col"]) < col
        ]
        if not candidates:
            continue

        candidates[-1]["date"] = orphan_date
        used_date_cols.add(col)

    return parsed


def fill_revision_dates(
    revisions: list[dict[str, object]],
    proposal_year: int,
    proposal_code: str,
    warnings: list[str],
) -> None:
    if not revisions:
        return

    known_indexes = [idx for idx, revision in enumerate(revisions) if revision["date"] is not None]

    if not known_indexes:
        base = date(proposal_year, 1, 1)
        for idx, revision in enumerate(revisions):
            revision["date"] = base + timedelta(days=idx)
        warnings.append(
            f"No revision dates in source, fallback applied: code={proposal_code}, base={base.isoformat()}"
        )
        return

    for idx in range(1, len(revisions)):
        if revisions[idx]["date"] is None and revisions[idx - 1]["date"] is not None:
            revisions[idx]["date"] = revisions[idx - 1]["date"] + timedelta(days=1)

    first_known = next((idx for idx, revision in enumerate(revisions) if revision["date"] is not None), None)
    if first_known is not None and first_known > 0:
        for idx in range(first_known - 1, -1, -1):
            revisions[idx]["date"] = revisions[idx + 1]["date"] - timedelta(days=1)

    for idx, revision in enumerate(revisions):
        if revision["date"] is None:
            anchor = revisions[idx - 1]["date"] if idx > 0 else date(proposal_year, 1, 1)
            revision["date"] = anchor + timedelta(days=1)


def extract_records(
    xlsx_path: Path,
) -> tuple[
    list[CustomerRecord],
    list[ProposalRecord],
    list[ProposalRevisionRecord],
    list[str],
    dict[str, int],
    dict[str, int],
]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)

    customer_sheets = [sheet for sheet in workbook.sheetnames if sheet not in SKIP_SHEETS]
    slug_by_sheet: dict[str, str] = {}
    used_slugs: set[str] = set()

    customers: list[CustomerRecord] = []
    for sheet in customer_sheets:
        slug = unique_slug(slugify(sheet), used_slugs)
        slug_by_sheet[sheet] = slug
        customers.append(
            CustomerRecord(
                name=sheet,
                slug=slug,
                notes=f'Imported from legacy workbook "{xlsx_path.name}" (sheet "{sheet}").',
            )
        )

    proposals: list[ProposalRecord] = []
    proposal_revisions: list[ProposalRevisionRecord] = []
    warnings: list[str] = []
    proposals_by_sheet: dict[str, int] = {sheet: 0 for sheet in customer_sheets}
    revisions_by_sheet: dict[str, int] = {sheet: 0 for sheet in customer_sheets}

    for sheet in customer_sheets:
        worksheet = workbook[sheet]
        mapping = extract_sheet_mapping(worksheet)
        customer_slug = slug_by_sheet[sheet]

        for row in range(5, worksheet.max_row + 1):
            raw_code = normalize_str(worksheet.cell(row=row, column=2).value)
            if not raw_code:
                continue

            match = PROPOSAL_CODE_RE.match(raw_code)
            if not match:
                continue

            _, year_raw, seq_raw = match.groups()
            year = int(year_raw)
            seq_number = int(seq_raw)

            invitation_code = ""
            if mapping.invitation_col is not None:
                invitation_code = normalize_str(
                    worksheet.cell(row=row, column=mapping.invitation_col).value
                )

            description = ""
            if mapping.description_col is not None:
                description = normalize_str(
                    worksheet.cell(row=row, column=mapping.description_col).value
                )
            if not description:
                description = f"Legacy proposal {raw_code}"
                warnings.append(
                    f"Missing description auto-filled: sheet={sheet}, row={row}, code={raw_code}"
                )

            active_value = None
            if mapping.active_col is not None:
                active_value = parse_decimal(worksheet.cell(row=row, column=mapping.active_col).value)

            won_value = None
            if mapping.won_col is not None:
                won_value = parse_decimal(worksheet.cell(row=row, column=mapping.won_col).value)

            lost_value = None
            if mapping.lost_col is not None:
                lost_value = parse_decimal(worksheet.cell(row=row, column=mapping.lost_col).value)

            if won_value is not None and lost_value is not None:
                warnings.append(
                    f"Both won/lost populated, won prioritized: sheet={sheet}, row={row}, code={raw_code}"
                )

            if won_value is not None:
                status = "ganha"
                final_value = won_value
                outcome_reason = ""
            elif lost_value is not None:
                status = "perdida"
                final_value = lost_value
                outcome_reason = ""
            else:
                status = "enviada"
                final_value = None
                outcome_reason = ""

            row_revisions = parse_revisions_from_row(worksheet, row, mapping)

            if not row_revisions:
                fallback_value = pick_first(active_value, won_value, lost_value)
                row_revisions = [
                    {
                        "label": 0,
                        "value_col": mapping.active_col or 0,
                        "value": fallback_value,
                        "date": None,
                    }
                ]
                if fallback_value is not None:
                    warnings.append(
                        f"No explicit revisions found, synthetic R0 created: sheet={sheet}, row={row}, code={raw_code}"
                    )
                else:
                    warnings.append(
                        f"No revision values in source, synthetic R0 with null value created: sheet={sheet}, row={row}, code={raw_code}"
                    )

            row_revisions.sort(key=lambda item: int(item["value_col"]))

            fill_revision_dates(row_revisions, year, raw_code, warnings)

            for index, revision in enumerate(row_revisions):
                raw_value_after = revision["value"]
                raw_value_before = row_revisions[index - 1]["value"] if index > 0 else None
                value_after = Decimal(raw_value_after) if raw_value_after is not None else None  # type: ignore[arg-type]
                value_before = Decimal(raw_value_before) if raw_value_before is not None else None  # type: ignore[arg-type]
                revision_date = revision["date"]  # type: ignore[assignment]
                proposal_revisions.append(
                    ProposalRevisionRecord(
                        proposal_code=raw_code,
                        revision_number=index,
                        value_before_brl=decimal_to_csv(value_before),
                        value_after_brl=decimal_to_csv(value_after),
                        reason="",
                        scope_changes="",
                        discount_brl="",
                        discount_percent="",
                        notes="",
                        created_at=date_to_timestamp_csv(revision_date),
                        legacy_sheet=sheet,
                        legacy_row=row,
                        legacy_revision_label=f"REV.{revision['label']}",
                    )
                )

            revision_values = [
                Decimal(item["value"]) for item in row_revisions if item["value"] is not None
            ]  # type: ignore[arg-type]
            estimated_value = pick_first(
                active_value,
                revision_values[-1] if revision_values else None,
                won_value,
                lost_value,
            )

            revision_dates = [item["date"] for item in row_revisions]
            created_at_date = min(revision_dates)
            updated_at_date = max(revision_dates)

            proposals.append(
                ProposalRecord(
                    customer_slug=customer_slug,
                    code=raw_code,
                    seq_number=seq_number,
                    year=year,
                    invitation_code=invitation_code,
                    project_name=description,
                    scope_description=description,
                    status=status,
                    estimated_value_brl=decimal_to_csv(estimated_value),
                    final_value_brl=decimal_to_csv(final_value),
                    outcome_reason=outcome_reason,
                    created_at=date_to_timestamp_csv(created_at_date),
                    updated_at=date_to_timestamp_csv(updated_at_date),
                    legacy_sheet=sheet,
                    legacy_row=row,
                )
            )

            proposals_by_sheet[sheet] += 1
            revisions_by_sheet[sheet] += len(row_revisions)

    code_counter = Counter([proposal.code for proposal in proposals])
    duplicate_codes = sorted([code for code, amount in code_counter.items() if amount > 1])
    if duplicate_codes:
        warnings.append(f"Duplicate proposal codes found: {', '.join(duplicate_codes)}")

    revision_counter = Counter(
        [(revision.proposal_code, revision.revision_number) for revision in proposal_revisions]
    )
    duplicate_revisions = [
        f"{code}-R{revision_number}"
        for (code, revision_number), amount in revision_counter.items()
        if amount > 1
    ]
    if duplicate_revisions:
        warnings.append(f"Duplicate revisions found: {', '.join(sorted(duplicate_revisions))}")

    return (
        customers,
        proposals,
        proposal_revisions,
        warnings,
        proposals_by_sheet,
        revisions_by_sheet,
    )


def write_customers_csv(path: Path, records: list[CustomerRecord]) -> None:
    fieldnames = ["name", "slug", "status", "cnpj", "notes"]
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "name": record.name,
                    "slug": record.slug,
                    "status": record.status,
                    "cnpj": record.cnpj,
                    "notes": record.notes,
                }
            )


def write_proposals_csv(path: Path, records: list[ProposalRecord]) -> None:
    fieldnames = [
        "customer_slug",
        "code",
        "seq_number",
        "year",
        "invitation_code",
        "project_name",
        "scope_description",
        "status",
        "estimated_value_brl",
        "final_value_brl",
        "outcome_reason",
        "created_at",
        "updated_at",
        "legacy_sheet",
        "legacy_row",
    ]
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "customer_slug": record.customer_slug,
                    "code": record.code,
                    "seq_number": record.seq_number,
                    "year": record.year,
                    "invitation_code": record.invitation_code,
                    "project_name": record.project_name,
                    "scope_description": record.scope_description,
                    "status": record.status,
                    "estimated_value_brl": record.estimated_value_brl,
                    "final_value_brl": record.final_value_brl,
                    "outcome_reason": record.outcome_reason,
                    "created_at": record.created_at,
                    "updated_at": record.updated_at,
                    "legacy_sheet": record.legacy_sheet,
                    "legacy_row": record.legacy_row,
                }
            )


def write_revisions_csv(path: Path, records: list[ProposalRevisionRecord]) -> None:
    fieldnames = [
        "proposal_code",
        "revision_number",
        "value_before_brl",
        "value_after_brl",
        "reason",
        "scope_changes",
        "discount_brl",
        "discount_percent",
        "notes",
        "created_at",
        "legacy_sheet",
        "legacy_row",
        "legacy_revision_label",
    ]
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "proposal_code": record.proposal_code,
                    "revision_number": record.revision_number,
                    "value_before_brl": record.value_before_brl,
                    "value_after_brl": record.value_after_brl,
                    "reason": record.reason,
                    "scope_changes": record.scope_changes,
                    "discount_brl": record.discount_brl,
                    "discount_percent": record.discount_percent,
                    "notes": record.notes,
                    "created_at": record.created_at,
                    "legacy_sheet": record.legacy_sheet,
                    "legacy_row": record.legacy_row,
                    "legacy_revision_label": record.legacy_revision_label,
                }
            )


def write_sql(
    path: Path,
    customers_csv: Path,
    proposals_csv: Path,
    revisions_csv: Path,
) -> None:
    sql = f"""\\set ON_ERROR_STOP on
\\if :{{?created_by}}
\\else
\\echo 'Required parameter: -v created_by=<UUID>'
\\quit 1
\\endif

BEGIN;

CREATE TEMP TABLE stg_customers (
  name text,
  slug text,
  status text,
  cnpj text,
  notes text
);

CREATE TEMP TABLE stg_proposals (
  customer_slug text,
  code text,
  seq_number integer,
  year integer,
  invitation_code text,
  project_name text,
  scope_description text,
  status text,
  estimated_value_brl text,
  final_value_brl text,
  outcome_reason text,
  created_at text,
  updated_at text,
  legacy_sheet text,
  legacy_row integer
);

CREATE TEMP TABLE stg_proposal_revisions (
  proposal_code text,
  revision_number integer,
  value_before_brl text,
  value_after_brl text,
  reason text,
  scope_changes text,
  discount_brl text,
  discount_percent text,
  notes text,
  created_at text,
  legacy_sheet text,
  legacy_row integer,
  legacy_revision_label text
);

\\copy stg_customers FROM '{customers_csv.as_posix()}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8');
\\copy stg_proposals FROM '{proposals_csv.as_posix()}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8');
\\copy stg_proposal_revisions FROM '{revisions_csv.as_posix()}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8');

INSERT INTO customers (name, slug, cnpj, notes, status)
SELECT
  sc.name,
  sc.slug,
  NULLIF(sc.cnpj, ''),
  NULLIF(sc.notes, ''),
  sc.status::customer_status
FROM stg_customers sc
ON CONFLICT (slug) DO UPDATE
SET
  name = EXCLUDED.name,
  cnpj = COALESCE(NULLIF(EXCLUDED.cnpj, ''), customers.cnpj),
  notes = COALESCE(NULLIF(EXCLUDED.notes, ''), customers.notes),
  status = EXCLUDED.status,
  updated_at = NOW();

INSERT INTO proposals (
  customer_id,
  code,
  seq_number,
  year,
  invitation_code,
  project_name,
  scope_description,
  status,
  due_date,
  estimated_value_brl,
  final_value_brl,
  outcome_reason,
  created_by,
  created_at,
  updated_at
)
SELECT
  c.id,
  sp.code,
  sp.seq_number,
  sp.year,
  NULLIF(sp.invitation_code, ''),
  sp.project_name,
  sp.scope_description,
  sp.status::proposal_status,
  NULL,
  NULLIF(sp.estimated_value_brl, '')::numeric(14, 2),
  NULLIF(sp.final_value_brl, '')::numeric(14, 2),
  NULLIF(sp.outcome_reason, ''),
  :'created_by'::uuid,
  NULLIF(sp.created_at, '')::timestamptz,
  NULLIF(sp.updated_at, '')::timestamptz
FROM stg_proposals sp
JOIN customers c ON c.slug = sp.customer_slug
ON CONFLICT (code) DO UPDATE
SET
  customer_id = EXCLUDED.customer_id,
  seq_number = EXCLUDED.seq_number,
  year = EXCLUDED.year,
  invitation_code = EXCLUDED.invitation_code,
  project_name = EXCLUDED.project_name,
  scope_description = EXCLUDED.scope_description,
  status = EXCLUDED.status,
  due_date = EXCLUDED.due_date,
  estimated_value_brl = EXCLUDED.estimated_value_brl,
  final_value_brl = EXCLUDED.final_value_brl,
  outcome_reason = EXCLUDED.outcome_reason,
  created_at = EXCLUDED.created_at,
  updated_at = EXCLUDED.updated_at;

INSERT INTO proposal_revisions (
  proposal_id,
  revision_number,
  reason,
  scope_changes,
  discount_brl,
  discount_percent,
  value_before_brl,
  value_after_brl,
  notes,
  created_by,
  created_at
)
SELECT
  p.id,
  sr.revision_number,
  NULLIF(sr.reason, ''),
  NULLIF(sr.scope_changes, ''),
  NULLIF(sr.discount_brl, '')::numeric(14, 2),
  NULLIF(sr.discount_percent, '')::numeric(5, 2),
  NULLIF(sr.value_before_brl, '')::numeric(14, 2),
  NULLIF(sr.value_after_brl, '')::numeric(14, 2),
  NULLIF(sr.notes, ''),
  :'created_by'::uuid,
  NULLIF(sr.created_at, '')::timestamptz
FROM stg_proposal_revisions sr
JOIN proposals p ON p.code = sr.proposal_code
ON CONFLICT (proposal_id, revision_number) DO UPDATE
SET
  reason = EXCLUDED.reason,
  scope_changes = EXCLUDED.scope_changes,
  discount_brl = EXCLUDED.discount_brl,
  discount_percent = EXCLUDED.discount_percent,
  value_before_brl = EXCLUDED.value_before_brl,
  value_after_brl = EXCLUDED.value_after_brl,
  notes = EXCLUDED.notes,
  created_by = EXCLUDED.created_by,
  created_at = EXCLUDED.created_at;

WITH imported_customers AS (
  SELECT c.id
  FROM customers c
  JOIN stg_customers sc ON sc.slug = c.slug
),
next_seq AS (
  SELECT
    ic.id AS customer_id,
    COALESCE(MAX(p.seq_number) + 1, 1) AS next_seq
  FROM imported_customers ic
  LEFT JOIN proposals p ON p.customer_id = ic.id
  GROUP BY ic.id
)
INSERT INTO proposal_sequences (customer_id, next_seq)
SELECT ns.customer_id, ns.next_seq
FROM next_seq ns
ON CONFLICT (customer_id) DO UPDATE
SET
  next_seq = GREATEST(proposal_sequences.next_seq, EXCLUDED.next_seq),
  updated_at = NOW();

COMMIT;
"""
    path.write_text(sql, encoding="utf-8")


def write_summary(
    path: Path,
    input_path: Path,
    customers: list[CustomerRecord],
    proposals: list[ProposalRecord],
    revisions: list[ProposalRevisionRecord],
    warnings: list[str],
    proposals_by_sheet: dict[str, int],
    revisions_by_sheet: dict[str, int],
) -> None:
    status_counter = Counter([proposal.status for proposal in proposals])
    summary = {
        "input_file": input_path.as_posix(),
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "customers_total": len(customers),
        "customers_with_proposals": sum(1 for _, count in proposals_by_sheet.items() if count > 0),
        "proposals_total": len(proposals),
        "revisions_total": len(revisions),
        "proposals_by_status": dict(status_counter),
        "proposals_by_customer": proposals_by_sheet,
        "revisions_by_customer": revisions_by_sheet,
        "warnings": warnings,
    }
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not input_path.exists():
        raise SystemExit(f"File not found: {input_path}")

    customers_csv = output_dir / "customers_legacy.csv"
    proposals_csv = output_dir / "proposals_legacy.csv"
    revisions_csv = output_dir / "proposal_revisions_legacy.csv"
    sql_file = output_dir / "import_legacy.sql"
    summary_file = output_dir / "summary.json"

    (
        customers,
        proposals,
        revisions,
        warnings,
        proposals_by_sheet,
        revisions_by_sheet,
    ) = extract_records(input_path)

    write_customers_csv(customers_csv, customers)
    write_proposals_csv(proposals_csv, proposals)
    write_revisions_csv(revisions_csv, revisions)
    write_sql(sql_file, customers_csv, proposals_csv, revisions_csv)
    write_summary(
        summary_file,
        input_path,
        customers,
        proposals,
        revisions,
        warnings,
        proposals_by_sheet,
        revisions_by_sheet,
    )

    print(f"OK: {input_path}")
    print(f"- customers: {customers_csv}")
    print(f"- proposals: {proposals_csv}")
    print(f"- revisions: {revisions_csv}")
    print(f"- sql: {sql_file}")
    print(f"- summary: {summary_file}")
    print(f"- total customers: {len(customers)}")
    print(f"- total proposals: {len(proposals)}")
    print(f"- total revisions: {len(revisions)}")


if __name__ == "__main__":
    main()
